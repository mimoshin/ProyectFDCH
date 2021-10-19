from openpyxl import load_workbook

DICT = ['A','B','C','D','E','F','G','H','J','K','L','M','N',
              'O','P','Q','R','S','T','U','V','W','X','Y','Z']

class Person():
    def __init__(self,f_name,s_name,fl_name,sl_name,gender):
        self.f_name = f_name
        self.s_name = s_name
        self.fl_name = fl_name
        self.sl_name = sl_name
        self.gender = gender
    
    def __str__(self):
        return "%s %s %s %s %s" % (self.f_name,self.s_name,self.fl_name,self.sl_name,self.gender)
        
def load ():
    ex = load_workbook('libro1.xlsx')
    hoja2 = ex['Hoja1']
    max_column, min_column =  hoja2.max_column, hoja2.min_column
    max_row, min_row = hoja2.max_row, hoja2.min_row
    person_list = []

    for row in range(min_column+1,max_row):
        indice = [DICT[0]+str(row),DICT[1]+str(row)]
        value = str(hoja2[indice[0]].value).split(' ')
    
        texto = ''
        if value.__len__() > 2:
            for x in value[1:]:
                texto += x+' '
            hoja2[indice[1]] = texto
        elif len(value) == 2:
            hoja2[indice[1]] = value[1]
    
        hoja2[indice[0]] = value[0]
    ex.save('doc.xlsx')

def load_2():
    ex = load_workbook('C:\\Users\\Franco\\Desktop\\Proyecto\\Codigo\\apps\\members\\doc.xlsx')
    hoja2 = ex['Hoja1']
    max_row, min_row = hoja2.max_row, hoja2.min_row
    person_list = []
    
    for row in range(min_row,max_row):
        indice = [DICT[0]+str(row),DICT[1]+str(row),DICT[2]+str(row),DICT[3]+str(row),DICT[4]+str(row)]
        value = [ hoja2[indice[0]].value, hoja2[indice[1]].value, hoja2[indice[2]].value, hoja2[indice[3]].value, hoja2[indice[4]].value]
        person_list.append(value)
    return person_list

"""
hoja2 = ex['Hoja2']
max_column, min_column =  hoja2.max_column, hoja2.min_column
max_row, min_row = hoja2.max_row, hoja2.min_row
person_list = []

for row in range(min_column+1,max_row):
    data = []
    for column in range(max_column):
        text = DICT[column]+str(row)
        value = str(hoja2[text].value)
        data.append(value.strip().upper())
    person_list.append(Person(data[0],data[1],data[2],data[3],data[4]))

for x in person_list:
    print(x)
"""